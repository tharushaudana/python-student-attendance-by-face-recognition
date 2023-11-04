import config as config
import openpyxl

students = {}

attempted_ids = []

def load():
    global students

    wb = openpyxl.load_workbook(config.student_data_path);

    ws = wb['Sheet1']

    properties = []

    i = 0;

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4):
        j = 0

        st_id = None

        for cell in row:
            if (i == 0):
                properties.append(cell.value)
            else:
                if (j == 0): 
                    st_id = cell.value
                    students[st_id] = {}
                else:
                    students[st_id][properties[j]] = cell.value

            j = j + 1

        i = i + 1

    wb.close()

    print(len(students), 'of students data loaded from file.')

def get(st_id):
    if (not st_id in students): return None
    return students[st_id]

def is_attempted(st_id):
    return st_id in attempted_ids

def mark_attempted(st_id):
    if (is_attempted(st_id)): return
    attempted_ids.append(st_id)