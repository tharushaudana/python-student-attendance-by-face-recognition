import openpyxl

workbook = openpyxl.load_workbook('assets/student_data.xlsx');

worksheet = workbook['Sheet1']

for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=4):
    
    data = []

    for cell in row:
        data.append(cell.value)

    print(data)