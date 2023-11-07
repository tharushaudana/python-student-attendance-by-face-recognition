import os
import config as config
import openpyxl

today_path = None
wb = None

def create_dir_if_not_exists():
    if os.path.exists(config.attendsheets_dir_path): return
    os.makedirs(config.attendsheets_dir_path)
    print("dir created", config.attendsheets_dir_path)

def set_today_path(date_str):
    global today_path
    today_path = config.attendsheets_dir_path + "/" + date_str + ".xlsx"

def open_wb():
    global wb

    if today_path == None: return

    if (os.path.exists(today_path)):
        wb = openpyxl.load_workbook(today_path)
    else:
        wb = openpyxl.Workbook()

def save_wb():
    if today_path == None: return
    wb.save(today_path)
    wb.close()

def create_wb_for_today(students, date_str):    
    create_dir_if_not_exists()

    set_today_path(date_str)

    if (os.path.exists(today_path)): return

    # create new...

    open_wb()

    classes = []

    header_added = False

    for st_id, st in students.items():
        c = st['class']

        if (c in classes):
            ws = wb[c]
        else:
            ws = wb.create_sheet(c)
            classes.append(c)
            header_added = False

        if (not header_added):
            h = list(st.keys())
            h.insert(0, 'id')
            h.append('present')
            h.append('time')
            
            ws.append(h)
            
            header_added = True

        r = list(st.values())
        r.insert(0, st_id)
        r.append('0')
        r.append('-')

        ws.append(r)


    save_wb()

    print("Excel file created for today.")

def find_preset_column_index(ws):
    header_values = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return header_values[0].index('present') + 1

def find_row_index_of_student(st_id, ws):
    i = 2

    print(ws.max_row)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1, values_only=True):
        print(row[0])
        if (row[0] == st_id): return i
        i = i + 1

    return None

def mark_attempt(st_id, class_name, time_str):
    open_wb()
    
    ws = wb[class_name]

    row = find_row_index_of_student(st_id, ws)
    col = find_preset_column_index(ws)

    print(row, col)

    ws.cell(row, col, value="1")
    ws.cell(row, col + 1, time_str)

    save_wb()